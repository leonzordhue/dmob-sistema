import os
from flask import Flask, render_template_string, request, redirect, url_for, send_file, flash
from flask_sqlalchemy import SQLAlchemy
from flask_login import LoginManager, login_user, logout_user, login_required, current_user, UserMixin
from werkzeug.security import generate_password_hash, check_password_hash
from openpyxl import Workbook
from io import BytesIO
from datetime import datetime
import sqlite3

app = Flask(__name__)
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///ramais.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['SECRET_KEY'] = 'chave-secreta-dmob-2024-porta-5021'

db = SQLAlchemy(app)
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'

# Models
class Usuario(db.Model, UserMixin):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(50), unique=True, nullable=False)
    password_hash = db.Column(db.String(255), nullable=False)
    nome = db.Column(db.String(100), nullable=False)
    nivel = db.Column(db.String(20), default='user')
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

class Ramal(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    codigo = db.Column(db.String(20))
    numero = db.Column(db.String(10))
    ramal_estrada = db.Column(db.String(255))
    municipio = db.Column(db.String(100))
    extensao_km = db.Column(db.Float)
    situacao = db.Column(db.String(50))
    revestimento = db.Column(db.String(50))
    inicio = db.Column(db.String(255))
    fim = db.Column(db.String(255))
    coordenada_inicio = db.Column(db.String(100))
    coordenada_fim = db.Column(db.String(100))
    created_by = db.Column(db.Integer)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

class LogRegistro(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    usuario = db.Column(db.String(100), nullable=False)
    acao = db.Column(db.String(50), nullable=False)
    descricao = db.Column(db.Text)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

@login_manager.user_loader
def load_user(user_id):
    return Usuario.query.get(int(user_id))

def registrar_log(usuario, acao, descricao):
    log = LogRegistro(
        usuario=usuario,
        acao=acao,
        descricao=descricao
    )
    db.session.add(log)
    db.session.commit()

# TEMPLATE DE LOGIN COMPLETO (SEM CREDENCIAIS)
LOGIN_TEMPLATE = '''
<!DOCTYPE html>
<html lang="pt-br">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>DMOB - Login</title>
    <style>
        body {
            font-family: Arial, sans-serif;
            margin: 0;
            padding: 0;
            background: linear-gradient(135deg, #2c3e50, #34495e);
            display: flex;
            justify-content: center;
            align-items: center;
            height: 100vh;
        }
        .login-container {
            background: white;
            padding: 2rem;
            border-radius: 10px;
            box-shadow: 0 10px 30px rgba(0,0,0,0.3);
            width: 100%;
            max-width: 400px;
        }
        .logo {
            text-align: center;
            margin-bottom: 1.5rem;
        }
        h2 {
            text-align: center;
            color: #2c3e50;
            margin-bottom: 1.5rem;
        }
        .form-group {
            margin-bottom: 1rem;
        }
        label {
            display: block;
            margin-bottom: 0.5rem;
            font-weight: bold;
            color: #2c3e50;
        }
        input[type="text"],
        input[type="password"] {
            width: 100%;
            padding: 0.75rem;
            border: 1px solid #ddd;
            border-radius: 4px;
            box-sizing: border-box;
            font-size: 1rem;
        }
        button {
            background: #3498db;
            color: white;
            padding: 0.75rem;
            border: none;
            border-radius: 4px;
            cursor: pointer;
            width: 100%;
            font-size: 1rem;
            margin-top: 1rem;
        }
        button:hover {
            background: #2980b9;
        }
        .error {
            background: #f8d7da;
            color: #721c24;
            padding: 0.75rem;
            border-radius: 4px;
            margin-bottom: 1rem;
        }
    </style>
</head>
<body>
    <div class="login-container">
        <div class="logo">
            <h1 style="color: #2c3e50; margin: 0;">🚧 DMOB</h1>
            <p style="color: #7f8c8d; margin: 0;">Departamento de Mobilidade</p>
        </div>

        <h2>🔐 Login</h2>

        {% if error %}
        <div class="error">{{ error }}</div>
        {% endif %}

        <form method="POST">
            <div class="form-group">
                <label for="username">👤 Usuário:</label>
                <input type="text" id="username" name="username" required placeholder="Digite seu usuário">
            </div>
            <div class="form-group">
                <label for="password">🔑 Senha:</label>
                <input type="password" id="password" name="password" required placeholder="Digite sua senha">
            </div>
            <button type="submit">Entrar no Sistema</button>
        </form>
    </div>
</body>
</html>
'''

# TEMPLATE DO DASHBOARD
DASHBOARD_TEMPLATE = '''
<!DOCTYPE html>
<html lang="pt-br">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>DMOB - Dashboard</title>
    <style>
        body {
            font-family: Arial, sans-serif;
            margin: 0;
            padding: 0;
            background: #f0f2f5;
        }
        .header {
            background: #2c3e50;
            color: white;
            padding: 1rem;
            text-align: center;
        }
        .nav {
            background: #34495e;
            padding: 1rem;
        }
        .nav a {
            color: white;
            text-decoration: none;
            margin: 0 1rem;
        }
        .container {
            max-width: 1200px;
            margin: 2rem auto;
            padding: 0 1rem;
        }
        .card {
            background: white;
            padding: 2rem;
            border-radius: 8px;
            box-shadow: 0 2px 10px rgba(0,0,0,0.1);
        }
        .stats-grid {
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(200px, 1fr));
            gap: 1rem;
            margin: 2rem 0;
        }
        .stat-card {
            background: #3498db;
            color: white;
            padding: 2rem;
            text-align: center;
            border-radius: 8px;
        }
        .actions-grid {
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(250px, 1fr));
            gap: 1rem;
            margin-top: 2rem;
        }
        .action-card {
            background: #27ae60;
            color: white;
            padding: 2rem;
            text-decoration: none;
            border-radius: 8px;
            text-align: center;
        }
        .action-card:hover {
            background: #219652;
        }
        .admin-only {
            background: #e74c3c !important;
        }
        .admin-only:hover {
            background: #c0392b !important;
        }
    </style>
</head>
<body>
    <div class="header">
        <h1>🚧 DMOB - Departamento de Mobilidade</h1>
        <p>Sistema de Cadastro de Ramais Rodoviários</p>
    </div>
    
    <div class="nav">
        <a href="/dashboard">📊 Dashboard</a>
        <a href="/cadastrar-ramal">📝 Cadastrar Ramal</a>
        <a href="/ramais">📋 Listar Ramais</a>
        <a href="/exportar-excel">📥 Exportar Excel</a>
        {% if current_user.nivel == 'admin' %}
        <a href="/usuarios">👥 Gerenciar Usuários</a>
        <a href="/logs">📋 Logs do Sistema</a>
        {% endif %}
        <a href="/logout" style="float: right;">🚪 Sair ({{ user_nome }})</a>
    </div>

    <div class="container">
        <div class="card">
            <h2>📊 Dashboard</h2>
            <p>Bem-vindo, <strong>{{ user_nome }}</strong>! ({{ 'Administrador' if current_user.nivel == 'admin' else 'Usuário' }})</p>
            
            <div class="stats-grid">
                <div class="stat-card">
                    <h3>📈 Total de Ramais</h3>
                    <p style="font-size: 2.5rem; margin: 0;">{{ total_ramais }}</p>
                </div>
                <div class="stat-card" style="background: #27ae60;">
                    <h3>👥 Total de Usuários</h3>
                    <p style="font-size: 2.5rem; margin: 0;">{{ total_usuarios }}</p>
                </div>
            </div>

            <div class="actions-grid">
                <a href="/cadastrar-ramal" class="action-card">
                    <h3>📝</h3>
                    <p>Cadastrar Novo Ramal</p>
                </a>
                <a href="/ramais" class="action-card" style="background: #e67e22;">
                    <h3>📋</h3>
                    <p>Ver Todos os Ramais</p>
                </a>
                <a href="/exportar-excel" class="action-card" style="background: #9b59b6;">
                    <h3>📊</h3>
                    <p>Exportar para Excel</p>
                </a>
                {% if current_user.nivel == 'admin' %}
                <a href="/usuarios" class="action-card admin-only">
                    <h3>👥</h3>
                    <p>Gerenciar Usuários</p>
                </a>
                <a href="/logs" class="action-card" style="background: #34495e;">
                    <h3>📋</h3>
                    <p>Logs do Sistema</p>
                </a>
                {% endif %}
            </div>
        </div>
    </div>
</body>
</html>
'''

@app.route('/')
def index():
    return redirect(url_for('login'))

@app.route('/login', methods=['GET', 'POST'])
def login():
    error = None
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        user = Usuario.query.filter_by(username=username).first()
        
        if user and check_password_hash(user.password_hash, password):
            login_user(user)
            registrar_log(user.username, 'LOGIN', f'Usuário {user.nome} fez login no sistema')
            return redirect(url_for('dashboard'))
        else:
            error = 'Usuário ou senha inválidos'
    
    return render_template_string(LOGIN_TEMPLATE, error=error)

@app.route('/logout')
@login_required
def logout():
    registrar_log(current_user.username, 'LOGOUT', f'Usuário {current_user.nome} fez logout do sistema')
    logout_user()
    return redirect(url_for('login'))

@app.route('/dashboard')
@login_required
def dashboard():
    total_ramais = Ramal.query.count()
    total_usuarios = Usuario.query.count()
    return render_template_string(DASHBOARD_TEMPLATE, 
                               total_ramais=total_ramais, 
                               total_usuarios=total_usuarios,
                               user_nome=current_user.nome,
                               current_user=current_user)

@app.route('/cadastrar-ramal', methods=['GET', 'POST'])
@login_required
def cadastrar_ramal():
    if request.method == 'POST':
        ramal = Ramal(
            codigo=request.form.get('codigo'),
            numero=request.form.get('numero'),
            ramal_estrada=request.form.get('ramal_estrada'),
            municipio=request.form.get('municipio'),
            extensao_km=float(request.form.get('extensao_km') or 0),
            situacao=request.form.get('situacao'),
            revestimento=request.form.get('revestimento'),
            inicio=request.form.get('inicio'),
            fim=request.form.get('fim'),
            coordenada_inicio=request.form.get('coordenada_inicio'),
            coordenada_fim=request.form.get('coordenada_fim'),
            created_by=current_user.id
        )
        
        db.session.add(ramal)
        db.session.commit()
        
        registrar_log(current_user.username, 'CADASTRO_RAMAL', 
                     f'Cadastrou ramal: {ramal.codigo} - {ramal.ramal_estrada}')
        
        return redirect(url_for('listar_ramais'))
    
    form_template = '''
    <!DOCTYPE html>
    <html>
    <head>
        <title>Cadastrar Ramal - DMOB</title>
        <style>
            body { font-family: Arial; margin: 0; background: #f0f2f5; }
            .header { background: #2c3e50; color: white; padding: 1rem; text-align: center; }
            .nav { background: #34495e; padding: 1rem; }
            .nav a { color: white; text-decoration: none; margin: 0 1rem; }
            .container { max-width: 800px; margin: 2rem auto; padding: 0 1rem; }
            .card { background: white; padding: 2rem; border-radius: 8px; box-shadow: 0 2px 10px rgba(0,0,0,0.1); }
            .form-group { margin-bottom: 1rem; }
            label { display: block; margin-bottom: 0.5rem; font-weight: bold; }
            input, select { width: 100%; padding: 0.75rem; border: 1px solid #ddd; border-radius: 4px; }
            button { background: #27ae60; color: white; padding: 0.75rem 1.5rem; border: none; border-radius: 4px; cursor: pointer; }
        </style>
    </head>
    <body>
        <div class="header">
            <h1>🚧 DMOB - Departamento de Mobilidade</h1>
            <p>Sistema de Cadastro de Ramais Rodoviários</p>
        </div>
        
        <div class="nav">
            <a href="/dashboard">📊 Dashboard</a>
            <a href="/cadastrar-ramal">📝 Cadastrar Ramal</a>
            <a href="/ramais">📋 Listar Ramais</a>
            <a href="/exportar-excel">📥 Exportar Excel</a>
            {% if current_user.nivel == 'admin' %}
            <a href="/usuarios">👥 Gerenciar Usuários</a>
            <a href="/logs">📋 Logs do Sistema</a>
            {% endif %}
            <a href="/logout" style="float: right;">🚪 Sair ({{ current_user.nome }})</a>
        </div>

        <div class="container">
            <div class="card">
                <h2>📝 Cadastrar Novo Ramal</h2>
                <form method="POST">
                    <div style="display: grid; grid-template-columns: 1fr 1fr; gap: 1rem;">
                        <div class="form-group">
                            <label>🔢 Código:</label>
                            <input type="text" name="codigo" required>
                        </div>
                        <div class="form-group">
                            <label>#️⃣ Número:</label>
                            <input type="text" name="numero">
                        </div>
                        <div class="form-group" style="grid-column: span 2;">
                            <label>🛣️ Ramal/Estrada:</label>
                            <input type="text" name="ramal_estrada" required>
                        </div>
                        <div class="form-group">
                            <label>🏙️ Município:</label>
                            <input type="text" name="municipio" required>
                        </div>
                        <div class="form-group">
                            <label>📏 Extensão (km):</label>
                            <input type="number" step="0.01" name="extensao_km">
                        </div>
                        <div class="form-group">
                            <label>📍 Início:</label>
                            <input type="text" name="inicio" placeholder="Local onde o ramal começa">
                        </div>
                        <div class="form-group">
                            <label>🏁 Fim:</label>
                            <input type="text" name="fim" placeholder="Local onde o ramal termina">
                        </div>
                        <div class="form-group">
                            <label>🗺️ Coordenada de Início:</label>
                            <input type="text" name="coordenada_inicio" placeholder="Ex: -3.123456, -60.123456">
                        </div>
                        <div class="form-group">
                            <label>🗺️ Coordenada de Fim:</label>
                            <input type="text" name="coordenada_fim" placeholder="Ex: -3.123456, -60.123456">
                        </div>
                        <div class="form-group">
                            <label>📊 Situação:</label>
                            <select name="situacao">
                                <option value="">Selecione...</option>
                                <option value="PAVIMENTADA">PAVIMENTADA</option>
                                <option value="LEITO NATURAL">LEITO NATURAL</option>
                                <option value="PLANEJADA">PLANEJADA</option>
                                <option value="EM OBRA">EM OBRA</option>
                                <option value="CONCLUÍDO">CONCLUÍDO</option>
                            </select>
                        </div>
                        <div class="form-group">
                            <label>🛠️ Revestimento:</label>
                            <input type="text" name="revestimento">
                        </div>
                    </div>
                    <button type="submit" style="margin-top: 1rem;">✅ Cadastrar Ramal</button>
                </form>
            </div>
        </div>
    </body>
    </html>
    '''
    return render_template_string(form_template, current_user=current_user)

@app.route('/ramais')
@login_required
def listar_ramais():
    ramais = Ramal.query.all()
    
    list_template = '''
    <!DOCTYPE html>
    <html>
    <head>
        <title>Ramais - DMOB</title>
        <style>
            body { font-family: Arial; margin: 0; background: #f0f2f5; }
            .header { background: #2c3e50; color: white; padding: 1rem; text-align: center; }
            .nav { background: #34495e; padding: 1rem; }
            .nav a { color: white; text-decoration: none; margin: 0 1rem; }
            .container { max-width: 1400px; margin: 2rem auto; padding: 0 1rem; }
            .card { background: white; padding: 2rem; border-radius: 8px; box-shadow: 0 2px 10px rgba(0,0,0,0.1); }
            table { width: 100%; border-collapse: collapse; margin-top: 1rem; font-size: 0.9rem; }
            th, td { padding: 0.75rem; text-align: left; border-bottom: 1px solid #ddd; }
            th { background: #34495e; color: white; position: sticky; top: 0; }
            .btn { padding: 0.5rem 1rem; border: none; border-radius: 4px; cursor: pointer; text-decoration: none; display: inline-block; margin: 0.2rem; }
            .btn-edit { background: #3498db; color: white; }
            .btn-delete { background: #e74c3c; color: white; }
            .btn:hover { opacity: 0.8; }
        </style>
    </head>
    <body>
        <div class="header">
            <h1>🚧 DMOB - Departamento de Mobilidade</h1>
            <p>Sistema de Cadastro de Ramais Rodoviários</p>
        </div>
        
        <div class="nav">
            <a href="/dashboard">📊 Dashboard</a>
            <a href="/cadastrar-ramal">📝 Cadastrar Ramal</a>
            <a href="/ramais">📋 Listar Ramais</a>
            <a href="/exportar-excel">📥 Exportar Excel</a>
            {% if current_user.nivel == 'admin' %}
            <a href="/usuarios">👥 Gerenciar Usuários</a>
            <a href="/logs">📋 Logs do Sistema</a>
            {% endif %}
            <a href="/logout" style="float: right;">🚪 Sair ({{ current_user.nome }})</a>
        </div>

        <div class="container">
            <div class="card">
                <h2>📋 Ramais Cadastrados</h2>
                <div style="overflow-x: auto;">
                    <table>
                        <thead>
                            <tr>
                                <th>Código</th>
                                <th>Ramal/Estrada</th>
                                <th>Município</th>
                                <th>Início</th>
                                <th>Fim</th>
                                <th>Coordenada Início</th>
                                <th>Coordenada Fim</th>
                                <th>Extensão (km)</th>
                                <th>Situação</th>
                                <th>Revestimento</th>
                                {% if current_user.nivel == 'admin' %}
                                <th>Ações</th>
                                {% endif %}
                            </tr>
                        </thead>
                        <tbody>
                            {% for ramal in ramais %}
                            <tr>
                                <td><strong>{{ ramal.codigo }}</strong></td>
                                <td>{{ ramal.ramal_estrada }}</td>
                                <td>{{ ramal.municipio }}</td>
                                <td>{{ ramal.inicio or '-' }}</td>
                                <td>{{ ramal.fim or '-' }}</td>
                                <td>{{ ramal.coordenada_inicio or '-' }}</td>
                                <td>{{ ramal.coordenada_fim or '-' }}</td>
                                <td>{{ ramal.extensao_km }}</td>
                                <td>{{ ramal.situacao }}</td>
                                <td>{{ ramal.revestimento }}</td>
                                {% if current_user.nivel == 'admin' %}
                                <td>
                                    <a href="/editar-ramal/{{ ramal.id }}" class="btn btn-edit">✏️ Editar</a>
                                    <a href="/excluir-ramal/{{ ramal.id }}" class="btn btn-delete" onclick="return confirm('Tem certeza que deseja excluir este ramal?')">🗑️ Excluir</a>
                                </td>
                                {% endif %}
                            </tr>
                            {% else %}
                            <tr>
                                <td colspan="{% if current_user.nivel == 'admin' %}11{% else %}10{% endif %}" style="text-align: center; padding: 2rem;">
                                    📭 Nenhum ramal cadastrado ainda.
                                </td>
                            </tr>
                            {% endfor %}
                        </tbody>
                    </table>
                </div>
            </div>
        </div>
    </body>
    </html>
    '''
    return render_template_string(list_template, ramais=ramais, current_user=current_user)

@app.route('/editar-ramal/<int:id>', methods=['GET', 'POST'])
@login_required
def editar_ramal(id):
    if current_user.nivel != 'admin':
        return redirect(url_for('listar_ramais'))
    
    ramal = Ramal.query.get_or_404(id)
    
    if request.method == 'POST':
        ramal.codigo = request.form.get('codigo')
        ramal.numero = request.form.get('numero')
        ramal.ramal_estrada = request.form.get('ramal_estrada')
        ramal.municipio = request.form.get('municipio')
        ramal.extensao_km = float(request.form.get('extensao_km') or 0)
        ramal.situacao = request.form.get('situacao')
        ramal.revestimento = request.form.get('revestimento')
        ramal.inicio = request.form.get('inicio')
        ramal.fim = request.form.get('fim')
        ramal.coordenada_inicio = request.form.get('coordenada_inicio')
        ramal.coordenada_fim = request.form.get('coordenada_fim')
        ramal.updated_at = datetime.utcnow()
        
        db.session.commit()
        
        registrar_log(current_user.username, 'EDIÇÃO_RAMAL', 
                     f'Editou ramal: {ramal.codigo} - {ramal.ramal_estrada}')
        
        return redirect(url_for('listar_ramais'))
    
    form_template = '''
    <!DOCTYPE html>
    <html>
    <head>
        <title>Editar Ramal - DMOB</title>
        <style>
            body { font-family: Arial; margin: 0; background: #f0f2f5; }
            .header { background: #2c3e50; color: white; padding: 1rem; text-align: center; }
            .nav { background: #34495e; padding: 1rem; }
            .nav a { color: white; text-decoration: none; margin: 0 1rem; }
            .container { max-width: 800px; margin: 2rem auto; padding: 0 1rem; }
            .card { background: white; padding: 2rem; border-radius: 8px; box-shadow: 0 2px 10px rgba(0,0,0,0.1); }
            .form-group { margin-bottom: 1rem; }
            label { display: block; margin-bottom: 0.5rem; font-weight: bold; }
            input, select { width: 100%; padding: 0.75rem; border: 1px solid #ddd; border-radius: 4px; }
            button { background: #3498db; color: white; padding: 0.75rem 1.5rem; border: none; border-radius: 4px; cursor: pointer; margin-right: 0.5rem; }
        </style>
    </head>
    <body>
        <div class="header">
            <h1>🚧 DMOB - Departamento de Mobilidade</h1>
            <p>Sistema de Cadastro de Ramais Rodoviários</p>
        </div>
        
        <div class="nav">
            <a href="/dashboard">📊 Dashboard</a>
            <a href="/cadastrar-ramal">📝 Cadastrar Ramal</a>
            <a href="/ramais">📋 Listar Ramais</a>
            <a href="/exportar-excel">📥 Exportar Excel</a>
            {% if current_user.nivel == 'admin' %}
            <a href="/usuarios">👥 Gerenciar Usuários</a>
            <a href="/logs">📋 Logs do Sistema</a>
            {% endif %}
            <a href="/logout" style="float: right;">🚪 Sair ({{ current_user.nome }})</a>
        </div>

        <div class="container">
            <div class="card">
                <h2>✏️ Editar Ramal</h2>
                <form method="POST">
                    <div style="display: grid; grid-template-columns: 1fr 1fr; gap: 1rem;">
                        <div class="form-group">
                            <label>🔢 Código:</label>
                            <input type="text" name="codigo" value="{{ ramal.codigo }}" required>
                        </div>
                        <div class="form-group">
                            <label>#️⃣ Número:</label>
                            <input type="text" name="numero" value="{{ ramal.numero or '' }}">
                        </div>
                        <div class="form-group" style="grid-column: span 2;">
                            <label>🛣️ Ramal/Estrada:</label>
                            <input type="text" name="ramal_estrada" value="{{ ramal.ramal_estrada }}" required>
                        </div>
                        <div class="form-group">
                            <label>🏙️ Município:</label>
                            <input type="text" name="municipio" value="{{ ramal.municipio }}" required>
                        </div>
                        <div class="form-group">
                            <label>📏 Extensão (km):</label>
                            <input type="number" step="0.01" name="extensao_km" value="{{ ramal.extensao_km or '' }}">
                        </div>
                        <div class="form-group">
                            <label>📍 Início:</label>
                            <input type="text" name="inicio" value="{{ ramal.inicio or '' }}" placeholder="Local onde o ramal começa">
                        </div>
                        <div class="form-group">
                            <label>🏁 Fim:</label>
                            <input type="text" name="fim" value="{{ ramal.fim or '' }}" placeholder="Local onde o ramal termina">
                        </div>
                        <div class="form-group">
                            <label>🗺️ Coordenada de Início:</label>
                            <input type="text" name="coordenada_inicio" value="{{ ramal.coordenada_inicio or '' }}" placeholder="Ex: -3.123456, -60.123456">
                        </div>
                        <div class="form-group">
                            <label>🗺️ Coordenada de Fim:</label>
                            <input type="text" name="coordenada_fim" value="{{ ramal.coordenada_fim or '' }}" placeholder="Ex: -3.123456, -60.123456">
                        </div>
                        <div class="form-group">
                            <label>📊 Situação:</label>
                            <select name="situacao">
                                <option value="">Selecione...</option>
                                <option value="PAVIMENTADA" {% if ramal.situacao == 'PAVIMENTADA' %}selected{% endif %}>PAVIMENTADA</option>
                                <option value="LEITO NATURAL" {% if ramal.situacao == 'LEITO NATURAL' %}selected{% endif %}>LEITO NATURAL</option>
                                <option value="PLANEJADA" {% if ramal.situacao == 'PLANEJADA' %}selected{% endif %}>PLANEJADA</option>
                                <option value="EM OBRA" {% if ramal.situacao == 'EM OBRA' %}selected{% endif %}>EM OBRA</option>
                                <option value="CONCLUÍDO" {% if ramal.situacao == 'CONCLUÍDO' %}selected{% endif %}>CONCLUÍDO</option>
                            </select>
                        </div>
                        <div class="form-group">
                            <label>🛠️ Revestimento:</label>
                            <input type="text" name="revestimento" value="{{ ramal.revestimento or '' }}">
                        </div>
                    </div>
                    <div style="margin-top: 1rem;">
                        <button type="submit">💾 Salvar Alterações</button>
                        <a href="/ramais" style="background: #95a5a6; color: white; padding: 0.75rem 1.5rem; border-radius: 4px; text-decoration: none;">❌ Cancelar</a>
                    </div>
                </form>
            </div>
        </div>
    </body>
    </html>
    '''
    return render_template_string(form_template, ramal=ramal, current_user=current_user)

@app.route('/excluir-ramal/<int:id>')
@login_required
def excluir_ramal(id):
    if current_user.nivel != 'admin':
        return redirect(url_for('listar_ramais'))
    
    ramal = Ramal.query.get_or_404(id)
    
    registrar_log(current_user.username, 'EXCLUSÃO_RAMAL', 
                 f'Excluiu ramal: {ramal.codigo} - {ramal.ramal_estrada}')
    
    db.session.delete(ramal)
    db.session.commit()
    
    return redirect(url_for('listar_ramais'))

@app.route('/exportar-excel')
@login_required
def exportar_excel():
    ramais = Ramal.query.all()
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Ramais Cadastrados"
    
    headers = ['Código', 'Número', 'Ramal/Estrada', 'Município', 'Início', 'Fim', 
               'Coordenada Início', 'Coordenada Fim', 'Extensão (km)', 'Situação', 'Revestimento']
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    for row, ramal in enumerate(ramais, 2):
        ws.cell(row=row, column=1, value=ramal.codigo)
        ws.cell(row=row, column=2, value=ramal.numero)
        ws.cell(row=row, column=3, value=ramal.ramal_estrada)
        ws.cell(row=row, column=4, value=ramal.municipio)
        ws.cell(row=row, column=5, value=ramal.inicio or '')
        ws.cell(row=row, column=6, value=ramal.fim or '')
        ws.cell(row=row, column=7, value=ramal.coordenada_inicio or '')
        ws.cell(row=row, column=8, value=ramal.coordenada_fim or '')
        ws.cell(row=row, column=9, value=ramal.extensao_km)
        ws.cell(row=row, column=10, value=ramal.situacao)
        ws.cell(row=row, column=11, value=ramal.revestimento)
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    registrar_log(current_user.username, 'EXPORT_EXCEL', 'Exportou dados para Excel')
    
    return send_file(buffer, as_attachment=True, download_name='Ramais_Cadastrados.xlsx')

@app.route('/usuarios')
@login_required
def gerenciar_usuarios():
    if current_user.nivel != 'admin':
        return redirect(url_for('dashboard'))
    
    usuarios = Usuario.query.all()
    
    usuarios_template = '''
    <!DOCTYPE html>
    <html>
    <head>
        <title>Usuários - DMOB</title>
        <style>
            body { font-family: Arial; margin: 0; background: #f0f2f5; }
            .header { background: #2c3e50; color: white; padding: 1rem; text-align: center; }
            .nav { background: #34495e; padding: 1rem; }
            .nav a { color: white; text-decoration: none; margin: 0 1rem; }
            .container { max-width: 1000px; margin: 2rem auto; padding: 0 1rem; }
            .card { background: white; padding: 2rem; border-radius: 8px; box-shadow: 0 2px 10px rgba(0,0,0,0.1); }
            table { width: 100%; border-collapse: collapse; margin-top: 1rem; }
            th, td { padding: 1rem; text-align: left; border-bottom: 1px solid #ddd; }
            th { background: #34495e; color: white; }
            .btn { padding: 0.5rem 1rem; border: none; border-radius: 4px; cursor: pointer; text-decoration: none; display: inline-block; }
            .btn-delete { background: #e74c3c; color: white; }
            .admin-badge { background: #e74c3c; color: white; padding: 0.3rem 0.6rem; border-radius: 12px; font-size: 0.8rem; }
            .user-badge { background: #3498db; color: white; padding: 0.3rem 0.6rem; border-radius: 12px; font-size: 0.8rem; }
        </style>
    </head>
    <body>
        <div class="header">
            <h1>🚧 DMOB - Departamento de Mobilidade</h1>
            <p>Sistema de Cadastro de Ramais Rodoviários</p>
        </div>
        
        <div class="nav">
            <a href="/dashboard">📊 Dashboard</a>
            <a href="/cadastrar-ramal">📝 Cadastrar Ramal</a>
            <a href="/ramais">📋 Listar Ramais</a>
            <a href="/exportar-excel">📥 Exportar Excel</a>
            <a href="/usuarios">👥 Gerenciar Usuários</a>
            <a href="/logs">📋 Logs do Sistema</a>
            <a href="/logout" style="float: right;">🚪 Sair ({{ current_user.nome }})</a>
        </div>

        <div class="container">
            <div class="card">
                <h2>👥 Gerenciar Usuários</h2>
                <table>
                    <thead>
                        <tr>
                            <th>ID</th>
                            <th>Usuário</th>
                            <th>Nome</th>
                            <th>Nível</th>
                            <th>Data de Criação</th>
                            <th>Ações</th>
                        </tr>
                    </thead>
                    <tbody>
                        {% for usuario in usuarios %}
                        <tr>
                            <td>{{ usuario.id }}</td>
                            <td>{{ usuario.username }}</td>
                            <td>{{ usuario.nome }}</td>
                            <td>
                                {% if usuario.nivel == 'admin' %}
                                <span class="admin-badge">👑 Admin</span>
                                {% else %}
                                <span class="user-badge">👤 Usuário</span>
                                {% endif %}
                            </td>
                            <td>{{ usuario.created_at.strftime('%d/%m/%Y %H:%M') }}</td>
                            <td>
                                {% if usuario.id != current_user.id %}
                                <a href="/excluir-usuario/{{ usuario.id }}" class="btn btn-delete" onclick="return confirm('Tem certeza que deseja excluir este usuário?')">🗑️ Excluir</a>
                                {% else %}
                                <em style="color: #7f8c8d;">Usuário atual</em>
                                {% endif %}
                            </td>
                        </tr>
                        {% endfor %}
                    </tbody>
                </table>
            </div>
        </div>
    </body>
    </html>
    '''
    return render_template_string(usuarios_template, usuarios=usuarios, current_user=current_user)

@app.route('/excluir-usuario/<int:id>')
@login_required
def excluir_usuario(id):
    if current_user.nivel != 'admin':
        return redirect(url_for('dashboard'))
    
    if id == current_user.id:
        return redirect(url_for('gerenciar_usuarios'))
    
    usuario = Usuario.query.get_or_404(id)
    
    registrar_log(current_user.username, 'EXCLUSÃO_USUARIO', 
                 f'Excluiu usuário: {usuario.username} - {usuario.nome}')
    
    db.session.delete(usuario)
    db.session.commit()
    
    return redirect(url_for('gerenciar_usuarios'))

@app.route('/logs')
@login_required
def ver_logs():
    if current_user.nivel != 'admin':
        return redirect(url_for('dashboard'))
    
    logs = LogRegistro.query.order_by(LogRegistro.created_at.desc()).limit(100).all()
    
    logs_template = '''
    <!DOCTYPE html>
    <html>
    <head>
        <title>Logs - DMOB</title>
        <style>
            body { font-family: Arial; margin: 0; background: #f0f2f5; }
            .header { background: #2c3e50; color: white; padding: 1rem; text-align: center; }
            .nav { background: #34495e; padding: 1rem; }
            .nav a { color: white; text-decoration: none; margin: 0 1rem; }
            .container { max-width: 1200px; margin: 2rem auto; padding: 0 1rem; }
            .card { background: white; padding: 2rem; border-radius: 8px; box-shadow: 0 2px 10px rgba(0,0,0,0.1); }
            table { width: 100%; border-collapse: collapse; margin-top: 1rem; font-size: 0.9rem; }
            th, td { padding: 0.75rem; text-align: left; border-bottom: 1px solid #ddd; }
            th { background: #34495e; color: white; }
            .log-login { background: #d4edda; }
            .log-logout { background: #fff3cd; }
            .log-cadastro { background: #d1ecf1; }
            .log-edicao { background: #e2e3e5; }
            .log-exclusao { background: #f8d7da; }
        </style>
    </head>
    <body>
        <div class="header">
            <h1>🚧 DMOB - Departamento de Mobilidade</h1>
            <p>Sistema de Cadastro de Ramais Rodoviários</p>
        </div>
        
        <div class="nav">
            <a href="/dashboard">📊 Dashboard</a>
            <a href="/cadastrar-ramal">📝 Cadastrar Ramal</a>
            <a href="/ramais">📋 Listar Ramais</a>
            <a href="/exportar-excel">📥 Exportar Excel</a>
            <a href="/usuarios">👥 Gerenciar Usuários</a>
            <a href="/logs">📋 Logs do Sistema</a>
            <a href="/logout" style="float: right;">🚪 Sair ({{ current_user.nome }})</a>
        </div>

        <div class="container">
            <div class="card">
                <h2>📋 Logs do Sistema</h2>
                <div style="overflow-x: auto;">
                    <table>
                        <thead>
                            <tr>
                                <th>Data/Hora</th>
                                <th>Usuário</th>
                                <th>Ação</th>
                                <th>Descrição</th>
                            </tr>
                        </thead>
                        <tbody>
                            {% for log in logs %}
                            <tr class="log-{{ log.acao.lower().split('_')[0] }}">
                                <td>{{ log.created_at.strftime('%d/%m/%Y %H:%M:%S') }}</td>
                                <td><strong>{{ log.usuario }}</strong></td>
                                <td>
                                    {% if log.acao == 'LOGIN' %}
                                    🔐 LOGIN
                                    {% elif log.acao == 'LOGOUT' %}
                                    🚪 LOGOUT
                                    {% elif log.acao == 'CADASTRO_RAMAL' %}
                                    📝 CADASTRO RAMAL
                                    {% elif log.acao == 'EDIÇÃO_RAMAL' %}
                                    ✏️ EDIÇÃO RAMAL
                                    {% elif log.acao == 'EXCLUSÃO_RAMAL' %}
                                    🗑️ EXCLUSÃO RAMAL
                                    {% elif log.acao == 'EXCLUSÃO_USUARIO' %}
                                    👥 EXCLUSÃO USUÁRIO
                                    {% else %}
                                    {{ log.acao }}
                                    {% endif %}
                                </td>
                                <td>{{ log.descricao }}</td>
                            </tr>
                            {% else %}
                            <tr>
                                <td colspan="4" style="text-align: center; padding: 2rem;">
                                    📭 Nenhum log registrado.
                                </td>
                            </tr>
                            {% endfor %}
                        </tbody>
                    </table>
                </div>
            </div>
        </div>
    </body>
    </html>
    '''
    return render_template_string(logs_template, logs=logs, current_user=current_user)

def init_db():
    """Inicializa o banco de dados"""
    with app.app_context():
        db.create_all()
        
        # Verificar se os usuários já existem
        if not Usuario.query.first():
            usuarios_iniciais = [
                {'username': 'admin', 'nome': 'Administrador DMOB', 'nivel': 'admin'},
                {'username': 'peneto', 'nome': 'Peneto', 'nivel': 'admin'},
                {'username': 'pedro.paiva', 'nome': 'Pedro Paiva', 'nivel': 'admin'},
                {'username': 'debora.binda', 'nome': 'Débora Binda', 'nivel': 'user'},
                {'username': 'kamilly.felipe', 'nome': 'Kamilly Felipe', 'nivel': 'user'}
            ]
            
            for user_data in usuarios_iniciais:
                usuario = Usuario(
                    username=user_data['username'],
                    password_hash=generate_password_hash('123456789'),
                    nome=user_data['nome'],
                    nivel=user_data['nivel']
                )
                db.session.add(usuario)
            
            db.session.commit()
            print('✅ Usuários criados: admin, peneto, pedro.paiva, debora.binda, kamilly.felipe')
            print('🔑 Senha para todos: 123456789')

if __name__ == '__main__':
    init_db()
    port = int(os.environ.get('PORT', 5021))
    print(f'🚀 Servidor iniciado na porta {port}: http://localhost:{port}')
    app.run(debug=True, port=port, host='0.0.0.0')